import csv
from importlib.resources import files
from typing import Dict, List, Tuple
from click.types import File
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import get_column_letter

from DataClass import Mutation, Sample


def mutations_dict(mutations: List[Mutation]) -> Tuple[List[str], List[Dict[str, str]]]:
    header = ['Sample']
    samples = {}
    for mutation in mutations:
        if mutation.found and mutation.name not in header:
            header.append(mutation.name)
        for sample in mutation.samples:
            if sample not in samples:
                samples[sample] = {'Sample': sample}
            samples[sample][mutation.name] = mutation.samples[sample]
    return header, samples.values()


def markers_dict(samples: List[Sample]) -> Tuple[List[str], List[Dict[str, str]]]:
    header = ['Sample', 'Marker mutations', 'Found mutations', 'Effect', 'Subtype', 'Papers']
    data = []
    for sample in samples:
        for marker in sample.markers:
            marker['Sample'] = sample.name
            data.append(marker)
    return header, data


def papers_dict(papers: List[Dict[str, str]]) -> Tuple[List[str], List[Dict[str, str]]]:
    header = ['Short name', 'Title', 'Authors', 'Year', 'Journal', 'Link', 'DOI']
    return header, papers


def write_csv(output_file: File, header: List[str], data: List[Dict[str, str]]) -> None:
    writer = csv.DictWriter(output_file, header, delimiter='\t', lineterminator='\n', extrasaction='ignore')
    writer.writeheader()
    writer.writerows(data)    


def get_workbook() -> Workbook:
    wb = load_workbook(files('data').joinpath('mutfinder_output.xlsm'), keep_vba=True)
    wb['Markers per sample']._pivots[0].cache.refreshOnLoad = True
    return wb


def save_workbook(wb: Workbook, save_path: str) -> None:
    wb.save(save_path)


def write_excel_sheet(wb: Workbook, sheet: str, header: List[str], data: List[Dict[str, str]]) -> Workbook:
    ws = wb[sheet]
    ws.append(header)
    for row, values in enumerate(data):
        for col, col_name in enumerate(header):
            ws.cell(row=row+2, column=col+1, value=values.get(col_name, ''))
    table = Table(displayName=f'{sheet}Table', ref=f'A1:{get_column_letter(len(header))}{len(data) + 1}')
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                                          showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    return wb
