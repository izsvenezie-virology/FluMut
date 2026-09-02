import csv
from collections import defaultdict
from io import TextIOWrapper

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import flumut
from flumut.core.analysis.models import Analysis
from flumut.core.globals import EXCEL_TEMPLATE
from flumut.core.logger import LOGGER
from flumut.core.options import OutputOptions
from flumut.flumutdb import DbVersion

TSV_data = list[dict[str, str]]


def write_outputs(analysis: Analysis, outputs: OutputOptions):
    """Write all analysis results to the requested output files.

    Extracts data from the completed Analysis object, writes tab-separated values
    to the provided file handles, and populates an Excel workbook with the same
    data using the built-in template.

    Args:
        analysis: The completed Analysis object.
        outputs: The requested reports, each an open file handle or None. The
            Excel handle's name determines its path and format (``.xlsx`` or ``.xlsm``).
    """
    markers_output = outputs.markers_output
    mutations_output = outputs.mutations_output
    literature_output = outputs.literature_output
    excel_output = outputs.excel_output

    if markers_output or excel_output:
        markers = get_markers_data(analysis)
        if markers_output:
            LOGGER.debug(f'Writing markers to {markers_output.name}')
            write_tsv(markers_output, markers)
    if mutations_output or excel_output:
        mutations = get_mutations_data(analysis)
        if mutations_output:
            LOGGER.debug(f'Writing mutations to {mutations_output.name}')
            write_tsv(mutations_output, mutations)
    if literature_output or excel_output:
        literature = get_literature_data(analysis)
        if literature_output:
            LOGGER.debug(f'Writing literature to {literature_output.name}')
            write_tsv(literature_output, literature)
    if excel_output:
        checks = get_checks_data(analysis)
        LOGGER.debug(f'Writing Excel report to {excel_output.name}')
        write_excel(excel_output, markers, mutations, literature, checks)  # type: ignore


def write_tsv(file: TextIOWrapper, values: TSV_data):
    """Write a list of row dictionaries to a tab-separated file.

    The column order follows the key order of the first row. Rows missing a
    key are written as empty strings. Does nothing and emits a warning if
    ``values`` is empty.

    Args:
        file: Open file handle for the output TSV file.
        values: List of row dictionaries to write.
    """
    if not values:
        LOGGER.warning(f'No data to write in file {file.name}')
        return
    header = values[0].keys()
    writer = csv.DictWriter(file, header, delimiter='\t', lineterminator='\n', extrasaction='ignore')
    writer.writeheader()
    writer.writerows(values)


def get_literature_data(analysis: Analysis) -> TSV_data:
    """Extract literature reference data from an analysis for output.

    Args:
        analysis: The completed Analysis object.

    Returns:
        A list of row dicts with keys: Short name, Title, Authors, Year,
        Journal, Link, DOI.
    """
    values = []

    for paper in sorted(analysis.literature, key=lambda paper: paper.short_name):
        paper_dict = {
            'Short name': paper.short_name,
            'Title': paper.title,
            'Authors': paper.authors,
            'Year': paper.year,
            'Journal': paper.journal,
            'Link': paper.url,
            'DOI': paper.doi,
        }
        values.append(paper_dict)
    return values


def get_markers_data(analysis: Analysis) -> TSV_data:
    """Extract detected marker data from an analysis for output.

    Groups evidences by (effect, subtype) and aggregates associated paper
    short names. One output row is produced per sample × marker × (effect, subtype)
    combination.

    Args:
        analysis: The completed Analysis object.

    Returns:
        A list of row dicts with keys: Sample, Marker, Mutations in your sample,
        Effect, Subtype, Literature.
    """
    values = []

    for sample in analysis.samples.values():
        for scan in sample.marker_scans:
            papers_collect = defaultdict(list)
            for evidence in scan.marker.evidences:
                effect_name = evidence.effect.name
                if evidence.host:
                    effect_name += f' in {evidence.host.name}'
                papers_collect[
                    (
                        effect_name,
                        evidence.subtype.name,
                    )
                ].append(evidence.paper.short_name)
            for (effect, subtype), papers in papers_collect.items():
                marker_evidence = {
                    'Sample': sample.id,
                    'Marker': scan.marker.name,
                    'Mutations in your sample': '; '.join([m.mutation.name for m in scan.detected_mutations]),
                    'Effect': effect,
                    'Subtype': subtype,
                    'Literature': '; '.join(papers),
                }
                values.append(marker_evidence)
    return values


def get_mutations_data(analysis: Analysis) -> TSV_data:
    """Extract per-sample mutation amino acid data from an analysis for output.

    Produces one row per sample with a column for each mutation detected across
    the entire analysis, filled with the observed amino acid(s) or left empty
    when the mutation was not scanned in that sample.

    Args:
        analysis: The completed Analysis object.

    Returns:
        A list of row dicts with ``'Sample'`` as the first key followed by one
        key per detected mutation, sorted by default position.
    """
    mutations = sorted(analysis.mutations)
    values = []

    for sample in analysis.samples.values():
        value = {'Sample': sample.id}
        mapping = {pos.mutation: pos for pos in sample.positions}
        for mutation in mutations:
            pos = mapping.get(mutation, None)
            if pos:
                value[mutation.name] = pos.ammino_acid
        values.append(value)
    return values


def get_checks_data(analysis: Analysis) -> TSV_data:
    """Extract quality-check messages from an analysis for output.

    Args:
        analysis: The completed Analysis object.

    Returns:
        A list of row dicts with a single ``'Checks'`` key containing the
        message for each issue detected across all samples.
    """
    values = []

    for sample in analysis.samples.values():
        for check in sample.checks:
            check_dict = {'Checks': check.message}
            values.append(check_dict)
    return values


def write_excel(output_file: TextIOWrapper, markers: TSV_data, mutations: TSV_data, literature: TSV_data, checks: TSV_data) -> None:
    """Write all analysis results to an Excel workbook using the built-in template.

    Loads the FluMut Excel template, writes version information to the Checks
    sheet, then populates the Mutations, Markers, Literature, and Checks sheets
    before saving to ``output_file``.

    Args:
        output_file: Open file handle whose name determines the output path and
            format (``.xlsx`` or ``.xlsm``).
        markers: Rows to write to the Markers sheet.
        mutations: Rows to write to the Mutations sheet.
        literature: Rows to write to the Literature sheet.
        checks: Rows to write to the Checks sheet.
    """
    wb = load_workbook(EXCEL_TEMPLATE, keep_vba=output_file.name.endswith('.xlsm'))

    ws = wb['Checks']
    ws.cell(row=1, column=4, value=flumut.__version__)
    ws.cell(row=1, column=7, value=str(DbVersion().get()))

    _write_excel_sheet(wb, 'Mutations', mutations)
    _write_excel_sheet(wb, 'Markers', markers)
    _write_excel_sheet(wb, 'Literature', literature)
    if checks:
        _write_excel_sheet(wb, 'Checks', checks)

    wb.save(output_file.name)


def _write_excel_sheet(wb: Workbook, sheet_name: str, values: TSV_data) -> None:
    """Populate one sheet of a workbook with tabular data and apply table formatting.

    Writes column headers from the first row's keys, fills subsequent rows with
    the corresponding values, and registers a styled Excel table over the data
    range. Does nothing if ``values`` is empty.

    Args:
        wb: The Workbook object to modify in-place.
        sheet_name: Name of the worksheet to populate.
        values: List of row dicts to write; column order follows the first row's keys.
    """
    if not values:
        LOGGER.warning(f'No data to write in sheet {sheet_name}')
        return
    ws = wb[sheet_name]
    header = list(values[0].keys())
    for col, col_name in enumerate(header):
        ws.cell(row=1, column=col + 1, value=col_name)
    for row, row_values in enumerate(values):
        for col, col_name in enumerate(header):
            ws.cell(row=row + 2, column=col + 1, value=row_values.get(col_name, ''))
    table = Table(displayName=f'{sheet_name}Table', ref=f'A1:{get_column_letter(len(header))}{len(values) + 1}')
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)
