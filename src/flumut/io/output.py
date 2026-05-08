import csv
from collections import defaultdict
from io import TextIOWrapper

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from flumut.analysis import Analysis
from flumut.core.globals import EXCEL_TEMPLATE

TSV_data = list[dict[str, str]]


def write_tsv(file: TextIOWrapper, values: TSV_data):
    header = values[0].keys()
    writer = csv.DictWriter(file, header, delimiter='\t', lineterminator='\n', extrasaction='ignore')
    writer.writeheader()
    writer.writerows(values)


def get_literature_data(analysis: Analysis) -> TSV_data:
    values = []

    for paper in analysis.literature:
        paper_dict = {
            'Short name': paper.short_name,
            'Title': paper.title,
            'Authors': paper.authors,
            'Year': paper.year,
            'Journal': paper.journal,
            'Link': paper.url,
            'DOI': paper.doi,
        }
        values.append(paper_dict)
    return values


def get_markers_data(analysis: Analysis) -> TSV_data:
    values = []

    for sample in analysis.samples.values():
        for scan in sample.marker_scans:
            papers_collect = defaultdict(list)
            for evidence in scan.marker.evidences:
                papers_collect[
                    (
                        evidence.effect.name,
                        evidence.subtype.name,
                        evidence.host.name if evidence.host else '',
                    )
                ].append(evidence.paper.short_name)
            for (effect, subtype, host), papers in papers_collect.items():
                marker_evidence = {
                    'Sample': sample.id,
                    'Marker': scan.marker.name if scan.marker.name else ', '.join([m.name for m in scan.marker.mutations]),
                    'Mutations in your sample': ', '.join([m.mutation.name for m in scan.detected_mutations]),
                    'Effect': effect,
                    'Subtype': subtype,
                    'Literature': ';'.join(papers),
                    'Host': host,
                }
                values.append(marker_evidence)
    return values


def get_mutations_data(analysis: Analysis) -> TSV_data:
    mutations = sorted(list(analysis.mutations), key=lambda mut: mut.default_position or 0)  # TODO: implement better sorting
    values = []

    for sample in analysis.samples.values():
        value = {'Sample': sample.id}
        mapping = {pos.mutation: pos for pos in sample.positions}
        for mutation in mutations:
            pos = mapping.get(mutation, None)
            if pos:
                value[mutation.name] = pos.ammino_acid
        values.append(value)
    return values


def write_excel(output_file: TextIOWrapper, markers: TSV_data, mutations: TSV_data, literature: TSV_data) -> None:
    """
    Write complete Excel output.

    :param `List[Sample]` samples: Samples to write in the output.
    :param `TextIOWrapper` output_file: Path for the output file.
    """

    wb = _open_workbook((output_file.name.endswith('.xlsm')))
    _write_excel_sheet(wb, 'Mutations', mutations)
    _write_excel_sheet(wb, 'Markers', markers)
    _write_excel_sheet(wb, 'Literature', literature)
    _save_workbook(wb, output_file)


def _open_workbook(keep_vba: bool) -> Workbook:
    """
    Open the template workbook.

    :param `bool` keep_vba: If `False` the VBA code in the template is discarded.
    :return `Workbook`: The opened workbook.
    """
    wb = load_workbook(EXCEL_TEMPLATE, keep_vba=keep_vba)
    return wb


def _write_excel_sheet(wb: Workbook, sheet_name: str, values: TSV_data) -> None:
    """
    Write an Excel sheet.

    :param `Workbook` wb: The workbook to modify.
    :param `str` sheet_name: The name of the sheet to modify.
    :param `List[str]` header: The header fields.
    :param `TSV_data` values: The list of values to write.
    """
    ws = wb[sheet_name]
    header = list(values[0].keys())
    ws.append(header)
    for row, row_values in enumerate(values):
        for col, col_name in enumerate(header):
            ws.cell(row=row + 2, column=col + 1, value=row_values.get(col_name, ''))
    table = Table(displayName=f'{sheet_name}Table', ref=f'A1:{get_column_letter(len(header))}{len(values) + 1}')
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)


def _save_workbook(wb: Workbook, output_file: TextIOWrapper) -> None:
    """
    Save the workbook in the specified path.

    :param `Workbook` wb: The workbook to save.
    :param `str` output_file: The path where save the workbook.
    """
    wb.save(output_file.name)
