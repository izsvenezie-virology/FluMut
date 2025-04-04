

import csv
from io import TextIOWrapper
from typing import Callable, Dict, List, Set, Tuple, Union

from importlib_resources import files
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import get_column_letter
from flumut.db_utility.db_data import literature_by_ids
from flumut.db_utility.db_models import Mutation
from flumut.sequence_utility.models import Sample

_outputs: Dict[str, TextIOWrapper] = {}
'''
    List of output files to save.
    Keys are the file type, must correspond to keys in `_output_types`.
'''


def set_output_file(type: str, file: Union[TextIOWrapper, str]) -> None:
    '''
    Store the output files to save.

    :param `str` type: Output type. Accepted values are `mutations_output`, `markers_output`, `literature_output`, `excel_output`.
    `None` is a valid value and is not stored.
    :param `TextIOWrapper`|`str` file: File path or opened file.
    '''
    if file is None:
        return
    _outputs[type] = file


def write_outputs(samples: List[Sample]) -> None:
    '''
    Writes all outputs saved.

    :param `List[Sample]` samples: Samples to write in the output.
    Samples must be completely analyzed.
    '''
    for type, file in _outputs.items():
        _outputs_type[type](samples, file)


def write_mutation_output(samples: List[Sample], output_file: TextIOWrapper) -> None:
    '''
    Write mutations text output.

    :param `List[Sample]` samples: Samples to write in the output.
    :param `TextIOWrapper` output_file: Opened output file.
    '''
    header, values = _prepare_mutation_output(samples)
    _write_tsv(output_file, header, values)


def write_marker_output(samples: List[Sample], output_file: TextIOWrapper) -> None:
    '''
    Write markers text output.

    :param `List[Sample]` samples: Samples to write in the output.
    :param `TextIOWrapper` output_file: Opened output file.
    '''
    header, values = _prepare_markers_output(samples)
    _write_tsv(output_file, header, values)


def write_literature_output(samples: List[Sample], output_file: TextIOWrapper) -> None:
    '''
    Write literature text output.

    :param `List[Sample]` samples: Samples to write in the output.
    :param `TextIOWrapper` output_file: Opened output file.
    '''
    header, values = _prepare_literature_output(samples)
    _write_tsv(output_file, header, values)


def write_excel_output(samples: List[Sample], output_file: TextIOWrapper) -> None:
    '''
    Write complete Excel output.

    :param `List[Sample]` samples: Samples to write in the output.
    :param `TextIOWrapper` output_file: Path for the output file.
    '''
    mutation_header, mutation_values = _prepare_mutation_output(samples)
    marker_header, marker_values = _prepare_markers_output(samples)
    literature_header, literature_values = _prepare_literature_output(samples)

    wb = _open_workbook((output_file.endswith('.xlsm')))
    _write_excel_sheet(wb, 'Mutations', mutation_header, mutation_values)
    _write_excel_sheet(wb, 'Markers', marker_header, marker_values)
    _write_excel_sheet(wb, 'Literature', literature_header, literature_values)
    _save_workbook(wb, output_file)


def _prepare_mutation_output(samples: List[Sample]) -> Tuple[List[str], List[Dict[str, str]]]:
    '''
    Prepare header and values for mutations results.

    :param `List[Sample]` samples: Samples to parse.
    '''
    mutations: Set[Mutation] = set()
    values: List[Dict[str, str]] = []
    for sample in samples:
        mutations.update(sample.mutations)
        aa_dict = sample.aas.copy()
        aa_dict['Sample'] = sample.name
        values.append(aa_dict)

    mutations = sorted(mutations, key=lambda x: (x.protein_name, list(x.mappings.values())[0].position))
    header = ['Sample'] + [mutation.name for mutation in mutations]
    return header, values


def _prepare_markers_output(samples: List[Sample]) -> Tuple[List[str], List[Dict[str, str]]]:
    '''
    Prepare header and values for markers results.

    :param `List[Sample]` samples: Samples to parse.
    '''
    header = ['Sample', 'Marker', 'Mutations in your sample', 'Effect', 'Subtype', 'Literature']
    values = []
    for sample in samples:
        for marker in sample.markers:
            value = marker.copy()
            value['Sample'] = sample.name
            values.append(value)
    return header, values


def _prepare_literature_output(samples: List[Sample]) -> Tuple[List[str], List[Dict[str, str]]]:
    '''
    Prepare header and values for literature results.

    :param `List[Sample]` samples: Samples to parse.
    '''
    header = ['Short name', 'Title', 'Authors', 'Year', 'Journal', 'Link', 'DOI']
    ids = set()
    for sample in samples:
        for marker in sample.markers:
            ids.update(marker['Literature'].split(';'))
    values = literature_by_ids(ids)
    return header, values


def _write_tsv(file: TextIOWrapper, header: List[str], values: List[Dict[str, str]]) -> None:
    '''
    Write header and values into a text file.

    :param `TextIOWrapper` file: The opened file to write.
    :param `List[str]` header: The header fields.
    :param `List[Dict[str, str]]` values: The list of values to write.
    Each element in the list is a single row.
    For each row, the dictionary must contain the field name as key and the field value as value.
    '''
    writer = csv.DictWriter(file, header, delimiter='\t', lineterminator='\n', extrasaction='ignore')
    writer.writeheader()
    writer.writerows(values)


def _open_workbook(keep_vba: bool) -> Workbook:
    '''
    Open the template workbook.

    :param `bool` keep_vba: If `False` the VBA code in the template is discarded.
    :return `Workbook`: The opened workbook.
    '''
    wb = load_workbook(files('flumut').joinpath('data/flumut_output.xlsm'), keep_vba=keep_vba)
    return wb


def _write_excel_sheet(wb: Workbook, sheet_name: str, header: List[str], values: List[Dict[str, str]]) -> None:
    '''
    Write an Excel sheet.

    :param `Workbook` wb: The workbook to modify.
    :param `str` sheet_name: The name of the sheet to modify.
    :param `List[str]` header: The header fields.
    :param `List[Dict[str, str]]` values: The list of values to write.
    '''
    ws = wb[sheet_name]
    ws.append(header)
    for row, row_values in enumerate(values):
        for col, col_name in enumerate(header):
            ws.cell(row=row+2, column=col+1, value=row_values.get(col_name, ''))
    table = Table(displayName=f'{sheet_name}Table', ref=f'A1:{get_column_letter(len(header))}{len(values) + 1}')
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                                          showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def _save_workbook(wb: Workbook, output_file: TextIOWrapper) -> None:
    '''
    Save the workbook in the specified path.

    :param `Workbook` wb: The workbook to save.
    :param `str` output_file: The path where save the workbook.
    '''
    output_file.close()
    wb.save(output_file.name)


_outputs_type: Dict[str, Callable] = {
    'mutations_output': write_mutation_output,
    'markers_output': write_marker_output,
    'literature_output': write_literature_output,
    'excel_output': write_excel_output
}
'''List of methods to call for each output type.'''
