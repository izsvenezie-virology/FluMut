"""End-to-end tests asserting properties that hold for any database version.

Which markers get reported depends entirely on FluMutDB content, so nothing
here asserts a marker name or a row count: a new marker or a new piece of
evidence must not turn these red. Exact output is pinned separately, in
``test_snapshot.py``.

Each test covers one property of the pipeline, so a failure name says what
broke. Assertions inside a test share one run rather than paying for their own.
"""

import re
import subprocess
import sys

from openpyxl import load_workbook

from tests.e2e.conftest import (
    LITERATURE_COLUMNS,
    MARKERS_COLUMNS,
    OUTPUTS,
    SINGLE_SAMPLE_FASTA,
    TSV_OUTPUTS,
)

#: The default declared by ``-n/--name-regex`` in ``flumut/cli.py``.
DEFAULT_NAME_REGEX = r'(?P<sample>.+)_(?P<segment>.+)'


def expected_sample_ids(fasta_path, name_regex: str = DEFAULT_NAME_REGEX) -> set[str]:
    """Derive the sample IDs FluMut should produce, straight from the FASTA headers.

    Mirrors ``parse_header``: a header that does not match the pattern becomes a
    sample ID on its own. This depends only on the input file and the regex,
    never on the database, which is what makes it a safe expectation.
    """
    pattern = re.compile(name_regex)
    samples = set()
    for line in fasta_path.read_text(encoding='utf-8').splitlines():
        if not line.startswith('>'):
            continue
        header = ' '.join(line[1:].split())
        match = pattern.match(header)
        samples.add(match.groupdict().get('sample', header) if match else header)
    return samples


def marker_pairs(result) -> set[tuple[str, str]]:
    """Return the (sample, marker) pairs reported, collapsing the per-evidence rows."""
    return {(row['Sample'], row['Marker']) for row in result.rows('markers')}


def test_run_produces_wellformed_outputs(default_run) -> None:
    """The pipeline completes and every output has the expected shape."""
    assert default_run.exception is None
    assert default_run.exit_code == 0

    for name in TSV_OUTPUTS:
        assert default_run.paths[name].exists(), f'{name} output was not written'
        assert default_run.paths[name].stat().st_size > 0, f'{name} output is empty'

    assert default_run.columns('markers') == MARKERS_COLUMNS
    assert default_run.columns('literature') == LITERATURE_COLUMNS

    mutation_columns = default_run.columns('mutations')
    assert mutation_columns[0] == 'Sample'
    assert len(mutation_columns) > 1, 'expected at least one mutation column'

    # Sample identity comes from the FASTA headers, never from the database.
    reported = {row['Sample'] for row in default_run.rows('mutations')}
    assert reported == expected_sample_ids(SINGLE_SAMPLE_FASTA)

    for row in default_run.rows('markers'):
        assert row['Marker'], 'marker name must never be empty'
        assert row['Effect'], f'missing effect for marker {row["Marker"]}'
        assert row['Subtype'], f'missing subtype for marker {row["Marker"]}'


def test_run_without_any_output_fails(flumut) -> None:
    """Analysing with no output requested writes nothing, so it must not report success."""
    assert flumut(outputs=()).exit_code != 0


def test_outputs_are_mutually_consistent(default_run) -> None:
    """Nothing referenced in one output is missing from the others."""
    samples = {row['Sample'] for row in default_run.rows('mutations')}
    marker_rows = default_run.rows('markers')

    assert {row['Sample'] for row in marker_rows} <= samples

    published = {row['Short name'] for row in default_run.rows('literature')}
    cited = {name for row in marker_rows for name in row['Literature'].split('; ') if name}
    assert cited <= published, f'cited but not listed in literature: {sorted(cited - published)}'

    columns = set(default_run.columns('mutations'))
    detected = {name for row in marker_rows for name in row['Mutations in your sample'].split('; ') if name}
    assert detected <= columns, f'reported but not tabulated: {sorted(detected - columns)}'


def test_excel_report_is_complete(default_run, flumut) -> None:
    """The workbook mirrors the TSVs, and is filled in even when no TSV is requested."""
    workbook = load_workbook(default_run.paths['excel'], keep_vba=True)
    assert {'Mutations', 'Markers', 'Literature', 'Checks'} <= set(workbook.sheetnames)

    for sheet_name, output in (('Markers', 'markers'), ('Literature', 'literature')):
        rows = sum(1 for row in workbook[sheet_name].iter_rows(min_row=2) if row[0].value not in (None, ''))
        assert rows == len(default_run.rows(output)), f'{sheet_name} sheet does not match {output}.tsv'

    # write_outputs must still build every dataset when only the Excel is asked for.
    excel_only = flumut(outputs=('excel',))
    assert excel_only.exception is None
    only_workbook = load_workbook(excel_only.paths['excel'], keep_vba=True)
    assert only_workbook['Markers'].max_row == workbook['Markers'].max_row


def test_relaxed_reports_a_superset_of_strict_markers(default_run, relaxed_run) -> None:
    """``--relaxed`` loosens the detection threshold, so it can only add markers."""
    strict, relaxed = marker_pairs(default_run), marker_pairs(relaxed_run)
    assert strict <= relaxed, f'lost under --relaxed: {sorted(strict - relaxed)}'


def test_output_is_reproducible_between_processes(tmp_path, tiny_fasta) -> None:
    """Two identical runs in fresh interpreters must produce identical files.

    Results are collected in sets of ORM instances, which iterate in an order
    derived from object hashes: stable within one process, but not between
    runs. Only a separate interpreter can catch that, so this shells out rather
    than reusing the in-process fixtures.
    """
    runs = []
    for attempt in ('first', 'second'):
        paths = {name: tmp_path / f'{attempt}-{name}.tsv' for name in TSV_OUTPUTS}
        args = [arg for name, path in paths.items() for arg in (OUTPUTS[name][0], str(path))]
        completed = subprocess.run(
            [sys.executable, '-m', 'flumut.cli', *args, str(tiny_fasta)],
            capture_output=True,
            text=True,
        )
        assert completed.returncode == 0, completed.stderr
        runs.append({name: path.read_text(encoding='utf-8') for name, path in paths.items()})

    for name in TSV_OUTPUTS:
        assert runs[0][name] == runs[1][name], f'{name}.tsv is not reproducible between processes'


def test_input_split_across_files_gives_the_same_result(flumut, tmp_path) -> None:
    """Splitting one sample's segments over two files must merge back into one sample."""
    records = SINGLE_SAMPLE_FASTA.read_text(encoding='utf-8').lstrip('>').split('\n>')
    first = tmp_path / 'first.fa'
    rest = tmp_path / 'rest.fa'
    first.write_text('>' + records[0].rstrip() + '\n', encoding='utf-8')
    rest.write_text('>' + '\n>'.join(records[1:]).rstrip() + '\n', encoding='utf-8')

    split = flumut(fasta=(first, rest), outputs=('mutations',))
    whole = flumut(outputs=('mutations',))

    assert split.exception is None
    assert split.text('mutations') == whole.text('mutations')
